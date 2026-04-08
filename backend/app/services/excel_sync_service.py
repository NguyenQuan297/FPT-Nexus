from __future__ import annotations

import asyncio
import json
import logging
import re
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from sqlalchemy import select

from app.core.config import settings
from app.db.session import AsyncSessionLocal
from app.models.import_batch import ImportBatch
from app.models.lead import Lead
from app.realtime.events import publish_event
from app.services.lead_display_utils import assignee_display_for_lead, build_username_display_map

log = logging.getLogger(__name__)

_lock = asyncio.Lock()
_last_run_ts: float = 0.0


def _sync_dir() -> Path:
    p = Path(settings.upload_dir) / "synced"
    p.mkdir(parents=True, exist_ok=True)
    return p


def _meta_path() -> Path:
    return _sync_dir() / "latest_sync.json"


def _template_info_path() -> Path:
    return _sync_dir() / "latest_template.json"


def _file_path() -> Path:
    return _sync_dir() / "leads_latest.xlsx"


async def get_latest_meta() -> Dict[str, Any]:
    mp = _meta_path()
    if not mp.exists():
        return {
            "status": "never_synced",
            "filename": None,
            "last_updated": None,
        }
    try:
        return json.loads(mp.read_text(encoding="utf-8"))
    except Exception:
        return {"status": "error", "filename": None, "last_updated": None}


async def _write_meta(meta: Dict[str, Any]) -> None:
    _meta_path().write_text(json.dumps(meta, ensure_ascii=True, indent=2), encoding="utf-8")


async def register_sync_template_upload(filename: str, raw: bytes) -> Dict[str, Any]:
    """
    Persist the original uploaded Excel as the canonical template for sync export.
    Export job will patch this file in-place so layout/styles stay unchanged.
    """
    ext = Path(filename).suffix.lower() or ".xlsx"
    if ext not in {".xlsx", ".xlsm"}:
        ext = ".xlsx"
    target = _sync_dir() / f"latest_template{ext}"
    target.write_bytes(raw)
    info = {
        "status": "ready",
        "filename": filename,
        "template_path": str(target),
        "saved_at": datetime.now(timezone.utc).isoformat(),
    }
    _template_info_path().write_text(json.dumps(info, ensure_ascii=True, indent=2), encoding="utf-8")
    return info


def _latest_template_path() -> Path | None:
    p = _template_info_path()
    if not p.exists():
        return None
    try:
        info = json.loads(p.read_text(encoding="utf-8"))
    except Exception:
        return None
    candidate = Path(str(info.get("template_path") or ""))
    if candidate.exists():
        return candidate
    return None


def _norm_text(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip()).lower()


def _phone_digits(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    text = re.sub(r"\.0$", "", text)
    digits = re.sub(r"\D", "", text)
    if len(digits) == 9:
        digits = "0" + digits
    return digits


_EXCEL_MAX_CELL_STR = 32000

# Fixed column order for snapshot export when no bundled template
CANONICAL_EXCEL_HEADERS: list[str] = [
    "Mã KH",
    "Tên khách hàng",
    "Số điện thoại",
    "Địa chỉ",
    "Nguồn khách hàng",
    "Người phụ trách",
    "Tình trạng cuộc gọi",
    "Tình trạng khách hàng",
    "Phân loại",
    "Ghi chú",
    "Trao đổi gần nhất",
    "Nhân viên phụ trách",
    "Mã số thuế",
    "Ngày tạo",
    "Ngày cập nhật",
    "Nhóm khách hàng",
    "Mô tả",
]


def _extra_field(ex: Dict[str, Any], *keys: str) -> str:
    for k in keys:
        v = ex.get(k)
        if v is not None and str(v).strip():
            return str(v).strip()
    return ""


def _dt_vn_full(dt: datetime | None) -> str:
    if dt is None:
        return ""
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=timezone.utc)
    return dt.astimezone(timezone.utc).strftime("%d/%m/%Y %H:%M:%S")


def _count_exchange_blocks(notes: str) -> int:
    if not (notes or "").strip():
        return 0
    txt = format_exchange_notes_for_excel_export(notes)
    return len([ln for ln in txt.split("\n") if ln.strip()])


def _one_line_to_img1(line: str) -> str:
    """Normalize one exchange line to: display_name HH:mm DD/MM/YYYY: body"""
    line = line.strip()
    if not line:
        return ""
    if re.match(r"^.+\s+\d{2}:\d{2}\s+\d{2}/\d{2}/\d{4}\s*:", line):
        return line

    m = re.match(r"^(.*)\s+(\d{2}/\d{2}/\d{4})\s+(\d{2}:\d{2}):\s*(.*)$", line)
    if m:
        body = m.group(4).strip().replace("\n", " ")
        suf = f": {body}" if body else ":"
        return f"{m.group(1).strip()} {m.group(3)} {m.group(2)}{suf}"

    m = re.match(r"^(.*)\s+(\d{2}/\d{2}/\d{4})\s+(\d{2}:\d{2})$", line)
    if m:
        return f"{m.group(1).strip()} {m.group(3)} {m.group(2)}:"

    m = re.match(r"^\[(\d{2}/\d{2}/\d{4} \d{2}:\d{2})\]\s+([^:]+):\s*(.*)$", line)
    if m:
        dt_part = m.group(1).split()
        if len(dt_part) >= 2:
            d0, t0 = dt_part[0], dt_part[1][:5]
            note = m.group(3).strip().replace("\n", " ")
            suf = f": {note}" if note else ":"
            return f"{m.group(2).strip()} {t0} {d0}{suf}"
    return line


def _normalize_exchange_block_to_img1(block: str) -> str:
    """One exchange block -> one canonical line."""
    block = block.strip()
    if not block:
        return ""
    lines = [ln.strip() for ln in block.split("\n") if ln.strip()]
    if len(lines) == 1:
        return _one_line_to_img1(lines[0])

    first, rest_lines = lines[0], lines[1:]
    rest = " ".join(rest_lines).replace("\n", " ")

    # Alt format: [DD/MM/YYYY HH:mm:ss] then "Name: note"
    if first.startswith("["):
        m = re.match(r"^\[(\d{2}/\d{2}/\d{4})\s+(\d{2}:\d{2}:\d{2})\]\s*$", first)
        m2 = re.match(r"^([^:]+):\s*(.*)$", rest)
        if m and m2:
            tshort = m.group(2)[:5]
            note = m2.group(2).strip()
            suf = f": {note}" if note else ":"
            return f"{m2.group(1).strip()} {tshort} {m.group(1)}{suf}"

    m = re.match(r"^(.*)\s+(\d{2}/\d{2}/\d{4})\s+(\d{2}:\d{2})$", first)
    if m:
        suf = f": {rest}" if rest else ":"
        return f"{m.group(1).strip()} {m.group(3)} {m.group(2)}{suf}"

    merged = f"{first} {rest}".strip()
    return _one_line_to_img1(merged)


def format_exchange_notes_for_excel_export(notes: str) -> str:
    """Multiple exchanges: one line each, newline-separated in the cell."""
    if not (notes or "").strip():
        return ""
    raw = notes.strip()
    blocks = re.split(r"\n\s*\n+", raw)
    out: list[str] = []
    for b in blocks:
        b = b.strip()
        if not b:
            continue
        out.append(_normalize_exchange_block_to_img1(b))
    return "\n".join(out)[:_EXCEL_MAX_CELL_STR]


def _load_workbook_template(path: Path):
    """Load workbook preserving rich text / styles as much as openpyxl allows."""
    return load_workbook(str(path), rich_text=True, keep_links=True, data_only=False)


def _merge_wrap_alignment(cell) -> None:
    """Keep cell fill/font; ensure text wraps for multi-line CRM content."""
    a = cell.alignment
    if a and (getattr(a, "wrap_text", None) or getattr(a, "wrapText", None)):
        return
    h = getattr(a, "horizontal", None) if a else None
    v = getattr(a, "vertical", None) if a else "top"
    cell.alignment = Alignment(
        wrap_text=True,
        horizontal=h if h is not None else "general",
        vertical=v if v is not None else "top",
        text_rotation=getattr(a, "text_rotation", 0) if a else 0,
        shrink_to_fit=getattr(a, "shrink_to_fit", None),
        indent=getattr(a, "indent", 0) if a else 0,
    )


def _bump_row_height_for_text(ws, row: int, text: str) -> None:
    t = str(text or "")
    lines = t.count("\n") + 1
    if len(t) > 120:
        lines = max(lines, min(80, int(len(t) / 70) + 2))
    target = min(409.0, max(15.0, 14.0 * lines * 1.1))
    rd = ws.row_dimensions[row]
    cur = rd.height
    if cur is None or cur < target:
        rd.height = target


def _exchange_text_for_excel(lead: Lead) -> str:
    """Latest exchange column: normalized name/time line + body."""
    n = (lead.notes or "").strip()
    if n:
        return format_exchange_notes_for_excel_export(n)
    ex = lead.extra or {}
    v = ex.get("Trao đổi gần nhất")
    if v is not None and str(v).strip():
        return format_exchange_notes_for_excel_export(str(v))
    if lead.last_contact_at:
        return lead.last_contact_at.astimezone(timezone.utc).strftime("%d/%m/%Y %H:%M")
    return ""


def _notes_for_excel(lead: Lead) -> str:
    n = (lead.notes or "").strip()
    return n[:_EXCEL_MAX_CELL_STR] if n else ""


def _last_contact_display(lead: Lead) -> str:
    if not lead.last_contact_at:
        return ""
    return lead.last_contact_at.astimezone(timezone.utc).strftime("%d/%m/%Y %H:%M")


def _status_for_excel(lead: Lead) -> str:
    if lead.status == "closed":
        return "Đóng"
    if lead.last_contact_at is not None or lead.status == "active":
        return "Đã liên hệ"
    if lead.status == "contacting":
        return "Đang liên hệ"
    if lead.status == "late":
        return "Trễ hạn"
    return "Chưa liên hệ"


def _status_cell_for_excel(lead: Lead) -> str:
    """Prefer call-status keys from extra if present, else derive from workflow."""
    ex = lead.extra or {}
    for k in ("Tình trạng cuộc gọi", "Tình trạng gọi điện"):
        v = ex.get(k)
        if v is not None and str(v).strip():
            return str(v).strip()
    return _status_for_excel(lead)


def _lead_row_canonical_values(lead: Lead, display_map: Dict[str, str]) -> list[str]:
    ex = lead.extra if isinstance(lead.extra, dict) else {}
    traodoi = format_exchange_notes_for_excel_export(lead.notes or "")
    assignee = assignee_display_for_lead(lead, display_map)
    return [
        (lead.external_id or "").strip(),
        (lead.name or "").strip(),
        (lead.phone or "").strip(),
        _extra_field(ex, "Địa chỉ", "Địa chỉ liên hệ"),
        (lead.source or "").strip(),
        assignee,
        _status_cell_for_excel(lead),
        _extra_field(ex, "Tình trạng khách hàng"),
        _extra_field(ex, "Phân loại"),
        _extra_field(ex, "Ghi chú"),
        traodoi,
        _extra_field(ex, "Nhân viên phụ trách", "Nhân viên cập nhật") or assignee,
        _extra_field(ex, "Mã số thuế", "Mã số thuế/CMND"),
        _dt_vn_full(lead.created_at),
        _dt_vn_full(lead.updated_at),
        _extra_field(ex, "Nhóm khách hàng"),
        _extra_field(ex, "Mô tả"),
    ]


def _resolve_template_path(uri: str | None) -> Path | None:
    if not uri or uri.startswith("s3://"):
        return None
    candidate = Path(uri)
    if candidate.exists():
        return candidate
    fallback = Path(settings.upload_dir) / candidate.name
    if fallback.exists():
        return fallback
    return None


def _pick_col(header_map: Dict[str, int], aliases: Iterable[str]) -> int | None:
    for alias in aliases:
        hit = header_map.get(_norm_text(alias))
        if hit is not None:
            return hit
    return None


def _find_header_row(ws) -> tuple[int, Dict[str, int]] | tuple[None, None]:
    for r in range(1, min(ws.max_row, 12) + 1):
        header_map: Dict[str, int] = {}
        for c in range(1, ws.max_column + 1):
            norm = _norm_text(ws.cell(row=r, column=c).value)
            if norm:
                header_map[norm] = c
        if not header_map:
            continue
        if any(
            key in header_map
            for key in (
                _norm_text("Mã KH"),
                _norm_text("Mã khách"),
                _norm_text("Mã khách hàng"),
                _norm_text("Tên khách hàng"),
                _norm_text("Ngày tạo"),
                _norm_text("Địa chỉ"),
                _norm_text("Số điện thoại"),
                _norm_text("Tình trạng cuộc gọi"),
                _norm_text("Tình trạng gọi điện"),
                _norm_text("Trao đổi gần nhất"),
                _norm_text("Ghi chú"),
            )
        ):
            return r, header_map
    return None, None


def _apply_updates_to_template(
    wb, rows: list[Lead], display_map: Dict[str, str]
) -> tuple[int, int]:
    by_external: Dict[str, Lead] = {}
    by_phone: Dict[str, Lead] = {}
    for lead in rows:
        ext = (lead.external_id or "").strip()
        if ext:
            by_external[ext] = lead
        phone = _phone_digits(lead.phone_normalized or lead.phone)
        if phone:
            by_phone[phone] = lead

    updated_cells = 0
    matched_rows = 0
    for ws in wb.worksheets:
        header_row, header_map = _find_header_row(ws)
        if not header_row or not header_map:
            continue

        col_external = _pick_col(
            header_map,
            ["Mã KH", "Mã khách", "Mã khách hàng", "Mã khách hàng CRM", "Mã đơn", "external_id"],
        )
        col_phone = _pick_col(
            header_map,
            [
                "Số điện thoại khách",
                "Số điện thoại",
                "Điện thoại phụ huynh",
                "Điện thoại",
                "Phone",
            ],
        )
        col_name = _pick_col(
            header_map,
            ["Tên khách hàng", "Họ tên khách", "Họ và tên", "Họ tên", "Tên Học Sinh", "Tên KH", "Name"],
        )
        col_dia_chi = _pick_col(header_map, ["Địa chỉ", "Địa chỉ liên hệ"])
        col_source = _pick_col(header_map, ["Nguồn khách hàng", "Nguồn KH", "Source", "Kênh"])
        col_assignee = _pick_col(header_map, ["Người phụ trách", "Assignee", "Assigned to"])
        col_status = _pick_col(
            header_map,
            ["Tình trạng cuộc gọi", "Tình trạng gọi điện", "Trạng thái gọi", "Status"],
        )
        col_tinh_trang_kh = _pick_col(header_map, ["Tình trạng khách hàng"])
        col_phan_loai = _pick_col(header_map, ["Phân loại"])
        col_nv_pt = _pick_col(header_map, ["Nhân viên phụ trách", "Nhân viên cập nhật"])
        col_nhom_kh = _pick_col(header_map, ["Nhóm khách hàng"])
        col_ngay_cap_nhat = _pick_col(
            header_map,
            ["Ngày cập nhật", "Thời gian cập nhật gần nhất"],
        )
        col_first_call = _pick_col(header_map, ["Lần đầu gọi điện", "Lần đầu gọi"])
        col_last_call = _pick_col(
            header_map,
            [
                "Lần cuối gọi điện",
                "Lần cuối gọi",
                "Lần gọi cuối",
                "Last contact at",
                "Lần liên hệ gần nhất",
                "Ngày liên hệ gần nhất",
                "Ngày giờ liên hệ",
            ],
        )
        col_so_lan_goi = _pick_col(header_map, ["Số lần gọi"])
        col_mo_ta = _pick_col(header_map, ["Mô tả", "Mô tả nhu cầu"])
        col_ma_thue = _pick_col(header_map, ["Mã số thuế/CMND", "Mã số thuế", "CMND", "Số CMND"])
        col_ngay_tao = _pick_col(header_map, ["Ngày tạo", "Created date", "Created at"])
        col_nguoi_tao = _pick_col(header_map, ["Người tạo"])
        col_khu_vuc = _pick_col(header_map, ["Khu vực", "Cơ sở"])
        col_chi_nhanh = _pick_col(header_map, ["Chi nhánh", "Branch", "Campus"])
        col_khoang_cach = _pick_col(header_map, ["Khoảng cách (km)", "Khoảng cách"])
        col_vi_tri = _pick_col(header_map, ["Vị trí khách hàng", "Vị trí"])
        col_exchange = _pick_col(
            header_map,
            ["Trao đổi gần nhất", "Nội dung trao đổi", "Trao đổi"],
        )
        col_notes = _pick_col(header_map, ["Ghi chú", "Notes", "Note", "Comment"])
        if not (col_external or col_phone):
            continue

        for r in range(header_row + 1, ws.max_row + 1):
            lead = None
            if col_external:
                ext = str(ws.cell(row=r, column=col_external).value or "").strip()
                if ext:
                    lead = by_external.get(ext)
            if lead is None and col_phone:
                phone = _phone_digits(ws.cell(row=r, column=col_phone).value)
                if phone:
                    lead = by_phone.get(phone)
            if lead is None:
                continue

            matched_rows += 1
            ex = lead.extra if isinstance(lead.extra, dict) else {}

            if col_name:
                nv = (lead.name or "").strip()
                cell = ws.cell(row=r, column=col_name)
                if str(cell.value or "").strip() != nv:
                    cell.value = nv or None
                    updated_cells += 1

            if col_dia_chi:
                dv = _extra_field(ex, "Địa chỉ", "Địa chỉ liên hệ")
                cell = ws.cell(row=r, column=col_dia_chi)
                if str(cell.value or "").strip() != dv:
                    cell.value = dv or None
                    updated_cells += 1

            if col_source:
                sv = (lead.source or "").strip()
                cell = ws.cell(row=r, column=col_source)
                if str(cell.value or "").strip() != sv:
                    cell.value = sv or None
                    updated_cells += 1

            if col_assignee:
                av = assignee_display_for_lead(lead, display_map)
                cell = ws.cell(row=r, column=col_assignee)
                if str(cell.value or "").strip() != av.strip():
                    cell.value = av or None
                    updated_cells += 1
                if av:
                    _merge_wrap_alignment(cell)

            if col_status:
                status_value = _status_cell_for_excel(lead)
                cell = ws.cell(row=r, column=col_status)
                if str(cell.value or "").strip() != status_value:
                    cell.value = status_value
                    updated_cells += 1

            if col_tinh_trang_kh:
                v = _extra_field(ex, "Tình trạng khách hàng")
                cell = ws.cell(row=r, column=col_tinh_trang_kh)
                if str(cell.value or "").strip() != v:
                    cell.value = v or None
                    updated_cells += 1

            if col_phan_loai:
                v = _extra_field(ex, "Phân loại")
                cell = ws.cell(row=r, column=col_phan_loai)
                if str(cell.value or "").strip() != v:
                    cell.value = v or None
                    updated_cells += 1

            if col_nv_pt:
                nv = _extra_field(ex, "Nhân viên phụ trách", "Nhân viên cập nhật") or assignee_display_for_lead(
                    lead, display_map
                )
                cell = ws.cell(row=r, column=col_nv_pt)
                if str(cell.value or "").strip() != nv.strip():
                    cell.value = nv or None
                    updated_cells += 1

            if col_nhom_kh:
                v = _extra_field(ex, "Nhóm khách hàng")
                cell = ws.cell(row=r, column=col_nhom_kh)
                if str(cell.value or "").strip() != v:
                    cell.value = v or None
                    updated_cells += 1

            if col_first_call:
                fv = _dt_vn_full(lead.contacted_at)
                cell = ws.cell(row=r, column=col_first_call)
                if str(cell.value or "").strip() != fv:
                    cell.value = fv or None
                    updated_cells += 1

            if col_last_call:
                lv = _dt_vn_full(lead.last_contact_at)
                cell = ws.cell(row=r, column=col_last_call)
                if str(cell.value or "").strip() != lv:
                    cell.value = lv or None
                    updated_cells += 1

            if col_so_lan_goi:
                n = _count_exchange_blocks(lead.notes or "")
                sv = str(n) if n else ""
                cell = ws.cell(row=r, column=col_so_lan_goi)
                if str(cell.value or "").strip() != sv:
                    cell.value = sv or None
                    updated_cells += 1

            if col_mo_ta:
                mv = _extra_field(ex, "Mô tả", "Mô tả nhu cầu")
                cell = ws.cell(row=r, column=col_mo_ta)
                if str(cell.value or "").strip() != mv:
                    cell.value = mv or None
                    updated_cells += 1

            if col_ma_thue:
                tv = _extra_field(ex, "Mã số thuế/CMND", "Mã số thuế", "CMND")
                cell = ws.cell(row=r, column=col_ma_thue)
                if str(cell.value or "").strip() != tv:
                    cell.value = tv or None
                    updated_cells += 1

            if col_ngay_tao:
                cv = _dt_vn_full(lead.created_at)
                cell = ws.cell(row=r, column=col_ngay_tao)
                if str(cell.value or "").strip() != cv:
                    cell.value = cv or None
                    updated_cells += 1

            if col_ngay_cap_nhat:
                uv = _dt_vn_full(lead.updated_at)
                cell = ws.cell(row=r, column=col_ngay_cap_nhat)
                if str(cell.value or "").strip() != uv:
                    cell.value = uv or None
                    updated_cells += 1

            if col_nguoi_tao:
                uv = _extra_field(ex, "Người tạo")
                cell = ws.cell(row=r, column=col_nguoi_tao)
                if str(cell.value or "").strip() != uv:
                    cell.value = uv or None
                    updated_cells += 1

            if col_khu_vuc:
                kv = _extra_field(ex, "Khu vực") or (lead.branch or "").strip()
                cell = ws.cell(row=r, column=col_khu_vuc)
                if str(cell.value or "").strip() != kv:
                    cell.value = kv or None
                    updated_cells += 1

            if col_chi_nhanh:
                bv = (lead.branch or "").strip()
                cell = ws.cell(row=r, column=col_chi_nhanh)
                if str(cell.value or "").strip() != bv:
                    cell.value = bv or None
                    updated_cells += 1

            if col_khoang_cach:
                gv = _extra_field(ex, "Khoảng cách (km)", "Khoảng cách")
                cell = ws.cell(row=r, column=col_khoang_cach)
                if str(cell.value or "").strip() != gv:
                    cell.value = gv or None
                    updated_cells += 1

            if col_vi_tri:
                vv = _extra_field(ex, "Vị trí khách hàng", "Vị trí")
                cell = ws.cell(row=r, column=col_vi_tri)
                if str(cell.value or "").strip() != vv:
                    cell.value = vv or None
                    updated_cells += 1

            exchange_value = _exchange_text_for_excel(lead) if col_exchange else ""

            if col_exchange:
                cell = ws.cell(row=r, column=col_exchange)
                if str(cell.value or "").strip() != (exchange_value or "").strip():
                    cell.value = exchange_value or None
                    updated_cells += 1
                if exchange_value:
                    _merge_wrap_alignment(cell)
                    _bump_row_height_for_text(ws, r, exchange_value)

            if col_notes:
                if col_exchange is not None and col_notes != col_exchange:
                    ghi = _extra_field(ex, "Ghi chú")
                    cell = ws.cell(row=r, column=col_notes)
                    if str(cell.value or "").strip() != ghi:
                        cell.value = ghi or None
                        updated_cells += 1
                    if ghi:
                        _merge_wrap_alignment(cell)
                else:
                    notes_value = _notes_for_excel(lead)
                    cell = ws.cell(row=r, column=col_notes)
                    if str(cell.value or "").strip() != (notes_value or "").strip():
                        cell.value = notes_value or None
                        updated_cells += 1
                    if notes_value:
                        _merge_wrap_alignment(cell)
                        _bump_row_height_for_text(ws, r, notes_value)

    return updated_cells, matched_rows


def _build_snapshot_workbook(rows: list[Lead], display_map: Dict[str, str]):
    """Fallback workbook when no template: 17 columns, yellow header row."""
    wb = Workbook()
    ws = wb.active
    ws.title = "danhsachkhachhang"
    hdr = CANONICAL_EXCEL_HEADERS
    ws.append(hdr)
    _hdr_fill = PatternFill(start_color="FFFAF200", end_color="FFFAF200", fill_type="solid")
    _hdr_font = Font(bold=True)
    for c in range(1, len(hdr) + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = _hdr_fill
        cell.font = _hdr_font
    for L in rows:
        vals = _lead_row_canonical_values(L, display_map)
        ws.append(vals)
        rr = ws.max_row
        for ci in range(1, len(hdr) + 1):
            if hdr[ci - 1] == "Trao đổi gần nhất":
                tc = ws.cell(row=rr, column=ci)
                if tc.value:
                    _merge_wrap_alignment(tc)
                    _bump_row_height_for_text(ws, rr, str(tc.value))
    return wb


async def generate_latest_excel(*, force: bool = False, min_interval_seconds: int = 45) -> Dict[str, Any]:
    """
    Build an up-to-date Excel snapshot from DB.
    Debounced to avoid generating file for every single row event.
    """
    global _last_run_ts
    now_ts = asyncio.get_running_loop().time()
    if not force and now_ts - _last_run_ts < min_interval_seconds:
        return await get_latest_meta()

    async with _lock:
        now_ts = asyncio.get_running_loop().time()
        if not force and now_ts - _last_run_ts < min_interval_seconds:
            return await get_latest_meta()

        async with AsyncSessionLocal() as db:
            rows = list((await db.execute(select(Lead).order_by(Lead.created_at.desc()))).scalars().all())
            latest_batch = (
                (await db.execute(select(ImportBatch).order_by(ImportBatch.created_at.desc()).limit(1)))
                .scalars()
                .first()
            )
            usernames = {(L.assigned_to or "").strip() for L in rows if (L.assigned_to or "").strip()}
            display_map = await build_username_display_map(db, usernames)

        source_template = _latest_template_path()
        if source_template is None:
            source_template = _resolve_template_path(latest_batch.storage_uri if latest_batch else None)
        template_error: str | None = None
        if source_template is not None:
            try:
                wb = _load_workbook_template(source_template)
                updated_cells, matched_rows = _apply_updates_to_template(wb, rows, display_map)
                mode = "template_patch"
            except Exception as e:
                log.warning("Template patch failed, fallback snapshot mode: %s", e, exc_info=True)
                template_error = str(e)[:800]
                wb = _build_snapshot_workbook(rows, display_map)
                updated_cells, matched_rows = 0, 0
                mode = "snapshot"
        else:
            wb = _build_snapshot_workbook(rows, display_map)
            updated_cells, matched_rows = 0, 0
            mode = "snapshot"

        out = _file_path()
        wb.save(out)
        stamp = datetime.now(timezone.utc).isoformat()
        meta = {
            "status": "synced",
            "filename": out.name,
            "path": str(out),
            "last_updated": stamp,
            "row_count": len(rows),
            "mode": mode,
            "template_source": str(source_template) if source_template else None,
            "matched_rows": matched_rows,
            "updated_cells": updated_cells,
            "template_error": template_error,
        }
        await _write_meta(meta)
        _last_run_ts = now_ts
        await publish_event("excel_sync.updated", meta)
        log.info(
            "Excel sync generated: rows=%s mode=%s matched=%s updated_cells=%s -> %s",
            len(rows),
            mode,
            matched_rows,
            updated_cells,
            out,
        )
        return meta

